import openpyxl
file_path = "D:\\desktop\\MTLS\\SU\\【】Methods in molecular life science\\Lab report\\ANNOVAR result\\Scer.variant_function_combine.xlsx"

def compare_sheets(control_sheet_name, experimental_sheet_name, difference_sheet_name, file_path):
    # 打开工作簿
    workbook = openpyxl.load_workbook(file_path)
    
    # 获取对照组、实验组和差异工作表
    control_sheet = workbook[control_sheet_name]
    experimental_sheet = workbook[experimental_sheet_name]
    
    # 创建差异工作表
    if difference_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=difference_sheet_name)
    difference_sheet = workbook[difference_sheet_name]
    
    # 清空差异工作表
    difference_sheet.delete_rows(1, difference_sheet.max_row)
    
    # 获取对照组和实验组的数据
    control_data = list(control_sheet.iter_rows(values_only=True))
    experimental_data = list(experimental_sheet.iter_rows(values_only=True))
    
    # 遍历对照组
    for control_row in control_data:
        different = True
        
        # 遍历实验组
        for experimental_row in experimental_data:
            # 比较前2列
            if control_row[1:3] == experimental_row[1:3]:
                different = False
                break
        
        # 如果有不同，则将数据复制到差异工作表
        if different:
            difference_sheet.append(control_row)
    
    # 保存工作簿
    workbook.save(file_path)

# 调用函数，传入对应的工作表名称和文件路径
compare_sheets("experiement", "control", "difference", file_path)

